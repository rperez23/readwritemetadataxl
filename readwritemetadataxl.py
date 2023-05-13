#! /usr/bin/python3

import os
import re
import sys
import openpyxl
import warnings
import subprocess
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font


warnings.simplefilter(action='ignore', category=UserWarning)

numeps   = 205

fnamecol           = 2
startrow           = 6
innhncol           = 4
innbuzzridcol      = 5
innairdatecol      = 6
inncontestantscol  = 10
deleterow          = 0

list38 = []
list39 = [128]    #start with row 128 since it will be deleted


#######Ask the user if these are masters or fc ready#######

print('')
assettype  = input('What type of media (fast or master): ')
seasonkeep = int(input('What season is this(38   or     39): '))
#print(assettype)

if (assettype != 'fast') and (assettype != 'master'): 
	print('   ~~~wrong asset~~~')
	sys.exit(0)
if (seasonkeep != 38) and (seasonkeep != 39):
	print('   ~~~wrong season~~~')
	sys.exit(0)

###########################################################


#open the input xl file, exit if it fails (The Database xl sheet)
xlinf    = 'TPIR Drew Database.xlsx'
xlintab  = 'TPIR Drew'
try:
    workbook = openpyxl.load_workbook(filename=xlinf,data_only=True)
except:
    print("  Cannot open",xlinf)
    sys.exit(0)
if not (xlintab in workbook.sheetnames):
    print(" ",xlintab,"not in",xlinf)
    workbook.close()
ws = workbook[xlintab]



#open the output xl file, exit if it fails
xloutf   = 'zMetadata_TPIR_DREW_2023.xlsx'
xlouttab = '1. Master Metadata'
try:
    wb2 = openpyxl.load_workbook(filename=xloutf)
except:
    print("  Cannot open",xloutf)
    sys.exit(0)
if not (xlouttab in wb2.sheetnames):
    print(" ",xlouttab,"not in",xloutf)
    wb2.close()
ws2 = wb2[xlouttab]

font = Font(name='Verdana',size=10)


#read  through the 'TPIR Drew Database' starting at row 6
#write to the      'zMetadata_TPIR_DREW_2023.xlsx' 
for i in range(0,numeps):

	row = startrow + i    

	housenum    = str(ws.cell(row,innhncol).value)           #get the housenum
	buzzrid     = str(ws.cell(row,innbuzzridcol).value)      #get the buzzrid
	airdate     = str(ws.cell(row,innairdatecol).value)      #get the airdate 

	

	#print('')
	#print('   ',buzzrid)
	if buzzrid == 'TPIR_EP5082_SR0038_YR2010_DC':
		deleterow = row
	
	contestants = str(ws.cell(row,inncontestantscol).value)  #get contestants
	contestants = contestants.replace(',',';')
	contestants = contestants.replace(',',';')
	contestants = contestants.replace('&','and')


	#split up buzzr id to get episode num, season
	#current name format of buzzrid: TPIR_EP4841_SR0038_YR2009_DC
	parts = buzzrid.split('_')                        

	epnum  = parts[1].replace('EP','')
	season = parts[2].replace('SR00','')

	#get the row number for each season and add to list
	if int(season) == 38:
		list38.append(row)
	else:
		list39.append(row)

	#set the movie name to ThePriceIsRight_s38_e4841_20230410.mov
	# TPIR_EP4841_SR0038_YR2009_DC -> ThePriceIsRight_s38_e4841_20230410.mov
	mov    = 'ThePriceIsRight_s' + season + '_e' + epnum + '_20230410.mov'

	#if we have a fast channel
	if assettype == 'fast':
		mov = mov.replace('.mov','.mxf')
		programversion = 'On-line Platform'
		capprefix = mov.split('.')[0] #caption prefix
	else:
		programversion = 'International'


	m = re.search('^\d+/\d+/\d+',airdate)

	#set the airdate to YYYY-DY-MO
	if m:
		parts   = airdate.split(' ')
		airdate = parts[0]
		parts   = airdate.split('/')
		mo      = parts[0].zfill(2)
		dy      = parts[1].zfill(2)
		yr      = str(int(parts[2]) + 2000)
		airdate = yr + '-' + mo + '-' + dy
	else:
		parts = airdate.split(' ')
		airdate = parts[0] 

	#print(season,':',buzzrid,':',mov,':',airdate)

	#write the data to the output xl file -> 

	ws2.cell(row=row,column=2).value  = mov
	ws2.cell(row=row,column=2).font = font

	ws2.cell(row=row,column=3).value  = "The Price is Right"
	ws2.cell(row=row,column=3).font = font

	ws2.cell(row=row,column=4).value  = "Television's longest-running game show, featuring host Drew Carey, where audience members try to win cash and prizes."
	ws2.cell(row=row,column=4).font = font

	ws2.cell(row=row,column=5).value  = "Episodic Television"
	ws2.cell(row=row,column=5).font = font

	ws2.cell(row=row,column=6).value  = "The Price is Right"
	ws2.cell(row=row,column=6).font = font

	ws2.cell(row=row,column=7).value  = "English"
	ws2.cell(row=row,column=7).font = font

	ws2.cell(row=row,column=8).value  = "English"
	ws2.cell(row=row,column=8).font = font

	ws2.cell(row=row,column=9).value  = airdate
	ws2.cell(row=row,column=9).font = font

	
	ws2.cell(row=row,column=10).value = int(season)
	ws2.cell(row=row,column=10).font = font

	ws2.cell(row=row,column=11).value = int(epnum)
	ws2.cell(row=row,column=11).font = font

	ws2.cell(row=row,column=13).value = "Non-scripted"
	ws2.cell(row=row,column=13).font = font

	ws2.cell(row=row,column=14).value = "Game Show"
	ws2.cell(row=row,column=14).font = font

	ws2.cell(row=row,column=16).value = "Drew Carey"
	ws2.cell(row=row,column=16).font = font

	ws2.cell(row=row,column=24).value = programversion
	ws2.cell(row=row,column=24).font = font

	ws2.cell(row=row,column=25).value = "Color"
	ws2.cell(row=row,column=25).font = font

	ws2.cell(row=row,column=32).value = buzzrid
	ws2.cell(row=row,column=32).font = font

	#Add the contestants (both cases)
	if contestants != 'None':
		ws2.cell(row=row,column=21).value = contestants 
		ws2.cell(row=row,column=21).font = font

	#Add the caption prefix
	if assettype == 'fast':
		ws2.cell(row=row,column=27).value = capprefix
		ws2.cell(row=row,column=27).font = font

		if housenum != 'None':
			ws2.cell(row=row,column=31).value = housenum
			ws2.cell(row=row,column=31).font = font


#close the source work book
workbook.close()

#Read through xl file ws2 (delete unwanted rows)
#first get list of buzzzrids to delete
#buzzridlist = ['TPIR_EP5082_SR0038_YR2010_DC']


if seasonkeep == 38:
	list39.reverse()
	deletelist = list39
	sr = 194

else:
	list38.reverse()
	deletelist = list38
	#sr = 

for delrow in deletelist:
	#print('delete : ',delrow)
	ws2.delete_rows(delrow,1)


if assettype == 'fast':

	#lscmd = 'aws s3 ls "s3://s3-fremantle-uk-or-1/fremantleuk/DMS UK/Media Files/T/ThePriceIsRight/Season38/FAST/"'
	#status = subprocess.run(lscmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, shell=True).stdout.strip("\n")
	#parts = status.split('\n')

	awscmd = 'aws s3 ls "s3://s3-fremantle-uk-or-1/fremantleuk/DMS UK/Media Files/T/ThePriceIsRight/Season' + str(seasonkeep) + '/FAST/"'
	#print(awscmd)
	status = subprocess.run(awscmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True, shell=True).stdout.strip("\n")
	parts = status.split('\n')
	#print(parts)

	mxflist = []

	for txt in parts:
		#2023-05-12 21:25:58 8035836140 ThePriceIsRight_s38_e5095_20230410.mxf

		m = re.search('\s+(ThePriceIsRight_s\d+_e\d+_\d+\.mxf)',txt)
		if m:
			#print(m.group(1))
			mxflist.append(m.group(1))

	#print(mxflist)

	for rownum in range(194,5,-1):

		episodename = str(ws2.cell(rownum,fnamecol).value)

		if not (episodename in mxflist):
			ws2.delete_rows(rownum,1)
		else:
			capprefix = str(ws2.cell(row=rownum,column=27).value)
			housenum  = str(ws2.cell(row=rownum,column=31).value)


			lncmd = 'ln ' + capprefix + '.scc ' + housenum + '.scc'
			print(lncmd)


wb2.save(xloutf)
wb2.close()
sys.exit(0)




